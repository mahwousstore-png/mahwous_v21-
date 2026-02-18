"""
engines/engine.py — محرك المطابقة v21
منطق: Fuzzy سريع → Gemini للغامض (62-96%) → تلقائي للواضح (97%+)
"""
import re, io, json, hashlib, sqlite3, time
from datetime import datetime
import pandas as pd
from rapidfuzz import fuzz, process as rf_process

try:
    from config import (REJECT_KEYWORDS, ALL_BRANDS, BRANDS_EN, BRANDS_AR,
                        KNOWN_BRANDS, WORD_REPLACEMENTS,
                        MATCH_THRESHOLD, AUTO_THRESHOLD, HIGH_CONFIDENCE,
                        REVIEW_THRESHOLD, PRICE_TOLERANCE,
                        TESTER_KEYWORDS, SET_KEYWORDS, GEMINI_API_KEYS,
                        DB_PATH, AI_BATCH_SIZE)
except Exception:
    REJECT_KEYWORDS = ["sample","عينة","عينه","decant","تقسيم","split","miniature"]
    ALL_BRANDS = KNOWN_BRANDS = BRANDS_EN = BRANDS_AR = []
    WORD_REPLACEMENTS = {}
    MATCH_THRESHOLD = 62; AUTO_THRESHOLD = 97; HIGH_CONFIDENCE = 92
    REVIEW_THRESHOLD = 75; PRICE_TOLERANCE = 10
    TESTER_KEYWORDS = ["tester","تستر"]; SET_KEYWORDS = ["set","طقم","مجموعة"]
    GEMINI_API_KEYS = []; DB_PATH = "mahwous.db"; AI_BATCH_SIZE = 12

import requests as _req

_GURL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"

# ══ مرادفات العطور ═══════════════════════════
_SYN = {
    "eau de parfum":"edp","او دو بارفان":"edp","أو دو بارفان":"edp",
    "او دي بارفان":"edp","بارفان":"edp","parfum":"edp","perfume":"edp",
    "eau de toilette":"edt","او دو تواليت":"edt","أو دو تواليت":"edt",
    "تواليت":"edt","toilette":"edt","toilet":"edt",
    "eau de cologne":"edc","كولون":"edc","cologne":"edc",
    "extrait de parfum":"extrait","parfum extrait":"extrait",
    "ديور":"dior","شانيل":"chanel","أرماني":"armani","فرساتشي":"versace",
    "غيرلان":"guerlain","توم فورد":"tom ford","لطافة":"lattafa","لطافه":"lattafa",
    "أجمل":"ajmal","رصاصي":"rasasi","أمواج":"amouage","كريد":"creed",
    "قوتشي":"gucci","برادا":"prada","هيرميس":"hermes","فالنتينو":"valentino",
    "كارتييه":"cartier","بولغاري":"bvlgari","ديزل":"diesel","سوفاج":"sauvage",
    "بلو":"bleu","وان ميليون":"1 million","إنفيكتوس":"invictus",
    "أفينتوس":"aventus","عود":"oud","مسك":"musk",
    " مل":" ml","ملي ":"ml ","ملي":"ml","مل":"ml",
    "أ":"ا","إ":"ا","آ":"ا","ة":"ه","ى":"ي","ؤ":"و","ئ":"ي",
}

# ══ SQLite Cache ══════════════════════════════
def _init_db():
    try:
        cn = sqlite3.connect(DB_PATH, check_same_thread=False)
        cn.execute("CREATE TABLE IF NOT EXISTS ai_cache(h TEXT PRIMARY KEY, v TEXT)")
        cn.commit(); cn.close()
    except Exception:
        pass

def _cget(k):
    try:
        cn = sqlite3.connect(DB_PATH, check_same_thread=False)
        r = cn.execute("SELECT v FROM ai_cache WHERE h=?", (k,)).fetchone()
        cn.close()
        return json.loads(r[0]) if r else None
    except Exception:
        return None

def _cset(k, v):
    try:
        cn = sqlite3.connect(DB_PATH, check_same_thread=False)
        cn.execute("INSERT OR REPLACE INTO ai_cache VALUES(?,?)", (k, json.dumps(v, ensure_ascii=False)))
        cn.commit(); cn.close()
    except Exception:
        pass

_init_db()

# ══ دوال أساسية ════════════════════════════
def read_file(f):
    """قراءة CSV/Excel مع دعم الترميزات العربية"""
    try:
        name = f.name.lower()
        if name.endswith('.csv'):
            for enc in ['utf-8', 'utf-8-sig', 'windows-1256', 'cp1256', 'latin-1']:
                try:
                    f.seek(0)
                    df = pd.read_csv(f, encoding=enc, on_bad_lines='skip')
                    if len(df) > 0: break
                except Exception:
                    continue
        elif name.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(f)
        else:
            return None, "صيغة غير مدعومة — CSV أو Excel فقط"
        df.columns = df.columns.str.strip()
        return df.dropna(how='all').reset_index(drop=True), None
    except Exception as e:
        return None, str(e)


def normalize(text):
    """تطبيع شامل للنص"""
    if not isinstance(text, str): return ""
    t = text.strip().lower()
    for k, v in WORD_REPLACEMENTS.items():
        t = t.replace(k.lower(), v)
    for k, v in _SYN.items():
        t = t.replace(k, v)
    t = re.sub(r'[^\w\s\u0600-\u06FF.]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()


def extract_size(text):
    if not isinstance(text, str): return 0.0
    m = re.findall(r'(\d+(?:\.\d+)?)\s*(?:ml|مل|ملي)', text.lower())
    return float(m[0]) if m else 0.0


def extract_brand(text):
    if not isinstance(text, str): return ""
    n = normalize(text)
    tl = text.lower()
    for b in ALL_BRANDS:
        nb = normalize(b)
        if nb and (nb in n or b.lower() in tl):
            return b
    return ""


def extract_type(text):
    if not isinstance(text, str): return ""
    n = normalize(text)
    if "extrait" in n: return "EXTRAIT"
    if "edp" in n: return "EDP"
    if "edt" in n: return "EDT"
    if "edc" in n: return "EDC"
    return ""


def is_sample(t):
    return isinstance(t, str) and any(k in t.lower() for k in REJECT_KEYWORDS)


def best_col(df, cands):
    """إيجاد أفضل عمود من قائمة مرشحين"""
    for c in cands:
        if c in df.columns: return c
    return df.columns[0] if len(df.columns) else ""


def get_price(row):
    for c in ["السعر", "سعر", "Price", "price", "PRICE"]:
        if c in row.index:
            try: return float(str(row[c]).replace(",", "").replace(" ", ""))
            except Exception: pass
    return 0.0


def get_id(row, col):
    if not col or col not in row.index: return ""
    v = str(row.get(col, ""))
    return "" if v in ("nan", "None", "") else v


# ══ فهرس المنافس (pre-normalized) ═════════════
class CompIndex:
    """يُبنى مرة واحدة — يسرّع المطابقة 5x"""
    def __init__(self, df, name_col, id_col, comp_name):
        self.comp_name  = comp_name
        self.raw_names  = df[name_col].fillna("").astype(str).tolist()
        self.norm_names = [normalize(n) for n in self.raw_names]
        self.brands     = [extract_brand(n) for n in self.raw_names]
        self.sizes      = [extract_size(n) for n in self.raw_names]
        self.types      = [extract_type(n) for n in self.raw_names]
        self.prices     = [get_price(row) for _, row in df.iterrows()]
        self.ids        = [get_id(row, id_col) for _, row in df.iterrows()]
        self._valid     = [i for i, n in enumerate(self.raw_names) if not is_sample(n) and n.strip()]

    def search(self, our_norm, our_br, our_sz, our_tp, top_n=5):
        if not self._valid: return []
        valid_norms = [self.norm_names[i] for i in self._valid]

        # مرحلة 1: بحث سريع بـ token_set_ratio
        fast = rf_process.extract(our_norm, valid_norms,
                                   scorer=fuzz.token_set_ratio,
                                   limit=min(25, len(valid_norms)))
        cands, seen = [], set()
        for _, fast_sc, vi in fast:
            if fast_sc < max(MATCH_THRESHOLD - 15, 40): continue
            idx  = self._valid[vi]
            name = self.raw_names[idx]
            if name in seen: continue

            c_br = self.brands[idx]
            c_sz = self.sizes[idx]
            c_tp = self.types[idx]

            # ── فلاتر صارمة ──
            if our_br and c_br and normalize(our_br) != normalize(c_br): continue
            if our_sz > 0 and c_sz > 0 and abs(our_sz - c_sz) > 30: continue
            if our_tp and c_tp and our_tp != c_tp:
                if our_sz > 0 and c_sz > 0 and abs(our_sz - c_sz) <= 3: continue

            # ── score مركّب دقيق ──
            n1, n2 = our_norm, self.norm_names[idx]
            s = (fuzz.token_sort_ratio(n1, n2) * 0.30
               + fuzz.token_set_ratio(n1, n2)  * 0.40
               + fuzz.partial_ratio(n1, n2)    * 0.30)

            if our_br and c_br:
                s += 8 if normalize(our_br) == normalize(c_br) else -22
            if our_sz > 0 and c_sz > 0:
                d = abs(our_sz - c_sz)
                s += 8 if d == 0 else (-5 if d <= 5 else -15 if d <= 20 else -28)
            if our_tp and c_tp and our_tp != c_tp:
                s -= 14

            score = round(max(0, min(100, s)), 1)
            if score < MATCH_THRESHOLD: continue
            seen.add(name)
            cands.append({
                "name": name, "score": score,
                "price": self.prices[idx], "product_id": self.ids[idx],
                "brand": c_br, "size": c_sz, "type": c_tp,
                "competitor": self.comp_name,
            })

        cands.sort(key=lambda x: x["score"], reverse=True)
        return cands[:top_n]


# ══ Gemini Batch ═════════════════════════════
def _ai_batch(batch):
    if not GEMINI_API_KEYS or not batch:
        return [0] * len(batch)

    ck = hashlib.md5(json.dumps(
        [{"o": x["our"], "c": [c["name"] for c in x["candidates"]]} for x in batch],
        ensure_ascii=False, sort_keys=True).encode()).hexdigest()
    cached = _cget(ck)
    if cached is not None:
        return cached

    lines = []
    for i, it in enumerate(batch):
        cands_text = "\n".join(
            f"  {j+1}. {c['name']} | {int(c.get('size',0))}ml | {c.get('type','?')} | {c.get('price',0):.0f}ر.س"
            for j, c in enumerate(it["candidates"])
        )
        lines.append(f"[{i+1}] منتجنا: «{it['our']}» ({it['price']:.0f}ر.س)\n{cands_text}")

    prompt = (
        "خبير عطور فاخرة. لكل منتج اختر رقم المرشح المطابق تماماً أو 0 إذا لا يوجد.\n"
        "شروط: ✓نفس الماركة ✓نفس الحجم (±5ml) ✓نفس EDP/EDT إذا مذكور\n\n"
        + "\n\n".join(lines)
        + f'\n\nJSON فقط: {{"results":[r1,r2,...,r{len(batch)}]}}'
    )
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0, "maxOutputTokens": 300, "topP": 1, "topK": 1}
    }
    for attempt in range(3):
        for key in GEMINI_API_KEYS:
            if not key: continue
            try:
                r = _req.post(f"{_GURL}?key={key}", json=payload, timeout=25)
                if r.status_code == 200:
                    txt = r.json()["candidates"][0]["content"]["parts"][0]["text"]
                    clean = re.sub(r'```json|```', '', txt).strip()
                    s = clean.find('{'); e = clean.rfind('}') + 1
                    if s >= 0 and e > s:
                        raw = json.loads(clean[s:e]).get("results", [])
                        out = []
                        for j, it in enumerate(batch):
                            n = raw[j] if j < len(raw) else 1
                            try: n = int(n)
                            except Exception: n = 1
                            if 1 <= n <= len(it["candidates"]): out.append(n - 1)
                            elif n == 0: out.append(-1)
                            else: out.append(0)
                        _cset(ck, out)
                        return out
                elif r.status_code == 429:
                    time.sleep(2 ** attempt)
            except Exception:
                continue
        time.sleep(1)
    return [0] * len(batch)


# ══ بناء صف النتيجة ════════════════════════
def _build_row(product, our_price, our_id, brand, size, ptype,
               best=None, src="", all_cands=None):
    sz_str = f"{int(size)}ml" if size else ""
    base = dict(المنتج=product, معرف_المنتج=our_id, السعر=our_price,
                الماركة=brand, الحجم=sz_str, النوع=ptype)
    if best is None:
        return {**base, **{
            "منتج_المنافس": "—", "معرف_المنافس": "", "سعر_المنافس": 0.0,
            "الفرق": 0.0, "الفرق_بالنسبة": 0.0, "نسبة_التطابق": 0.0,
            "القرار": "🔵 مفقود عند المنافس", "الخطورة": "",
            "المنافس": "", "مصدر_المطابقة": "—", "جميع_المرشحين": [],
        }}

    cp    = float(best.get("price") or 0)
    score = float(best.get("score") or 0)
    diff  = round(our_price - cp, 2) if our_price > 0 and cp > 0 else 0.0
    pct   = round(diff / cp * 100, 1) if cp > 0 else 0.0
    risk  = "🔴 عالي" if abs(diff) > 30 else ("🟡 متوسط" if abs(diff) > 10 else "🟢 منخفض")

    if abs(diff) <= PRICE_TOLERANCE:   dec = "✅ موافق عليها"
    elif diff > PRICE_TOLERANCE:       dec = "🔴 سعر أعلى"
    else:                              dec = "🟢 سعر أقل"

    if score < REVIEW_THRESHOLD and src not in ("auto", "gemini"):
        dec = "⚠️ مراجعة"

    src_label = {"auto": f"⚡ تلقائي({score:.0f}%)", "gemini": f"🤖 AI({score:.0f}%)"}.get(src, f"{score:.0f}%")

    return {**base, **{
        "منتج_المنافس": best["name"],
        "معرف_المنافس": best.get("product_id", ""),
        "سعر_المنافس":  cp,
        "الفرق":        diff,
        "الفرق_بالنسبة": pct,
        "نسبة_التطابق": score,
        "القرار":       dec,
        "الخطورة":      risk,
        "المنافس":      best.get("competitor", ""),
        "مصدر_المطابقة": src_label,
        "جميع_المرشحين": (all_cands or [best])[:5],
    }}


# ══ التحليل الكامل ════════════════════════════
def run_analysis(our_df, comp_dfs, progress_cb=None, use_ai=True):
    """
    المحرك الرئيسي — سريع + دقيق
    our_df: DataFrame ملف مهووس (يحتوي أعمدة: المنتج، السعر، معرف_المنتج)
    comp_dfs: {اسم: DataFrame} ملفات المنافسين
    progress_cb: دالة(0.0→1.0)
    """
    results = []
    our_name_col  = best_col(our_df, ["المنتج", "اسم المنتج", "Product", "Name", "name", "اسم"])
    our_price_col = best_col(our_df, ["السعر", "سعر", "Price", "price", "PRICE"])
    our_id_col    = best_col(our_df, ["معرف_المنتج", "no", "NO", "No", "معرف", "ID", "id", "SKU", "sku", "الكود"])

    # بناء الفهارس مرة واحدة
    indices = {}
    for cname, cdf in comp_dfs.items():
        cn_col = best_col(cdf, ["المنتج", "اسم المنتج", "Product", "Name", "name", "اسم"])
        ci_col = best_col(cdf, ["ID", "id", "معرف", "SKU", "sku", "الكود", "code", "no", "NO"])
        indices[cname] = CompIndex(cdf, cn_col, ci_col, cname)

    total   = len(our_df)
    pending = []

    def flush():
        if not pending: return
        idxs = _ai_batch(pending)
        for j, it in enumerate(pending):
            ci = idxs[j] if j < len(idxs) else 0
            if ci < 0:
                results.append(_build_row(it["product"], it["our_price"], it["our_id"],
                                           it["brand"], it["size"], it["ptype"], src="gemini"))
            else:
                best = it["candidates"][ci]
                results.append(_build_row(it["product"], it["our_price"], it["our_id"],
                                           it["brand"], it["size"], it["ptype"],
                                           best=best, src="gemini", all_cands=it["all_cands"]))
        pending.clear()

    for i, (_, row) in enumerate(our_df.iterrows()):
        product = str(row.get(our_name_col, "")).strip()
        if not product or is_sample(product):
            if progress_cb: progress_cb((i + 1) / total)
            continue

        try: our_price = float(str(row.get(our_price_col, 0)).replace(",", ""))
        except Exception: our_price = 0.0
        our_id = get_id(row, our_id_col)
        brand  = extract_brand(product)
        size   = extract_size(product)
        ptype  = extract_type(product)
        our_n  = normalize(product)

        # جمع المرشحين من كل المنافسين
        all_cands = []
        for idx_obj in indices.values():
            all_cands.extend(idx_obj.search(our_n, brand, size, ptype, top_n=5))

        if not all_cands:
            results.append(_build_row(product, our_price, our_id, brand, size, ptype))
            if progress_cb: progress_cb((i + 1) / total)
            continue

        all_cands.sort(key=lambda x: x["score"], reverse=True)
        best = all_cands[0]

        if best["score"] >= AUTO_THRESHOLD or not use_ai:
            results.append(_build_row(product, our_price, our_id, brand, size, ptype,
                                       best=best, src="auto", all_cands=all_cands))
        else:
            pending.append(dict(product=product, our_price=our_price, our_id=our_id,
                                brand=brand, size=size, ptype=ptype,
                                candidates=all_cands[:5], all_cands=all_cands,
                                our=product, price=our_price))
            if len(pending) >= AI_BATCH_SIZE:
                flush()

        if progress_cb: progress_cb((i + 1) / total)

    flush()
    return pd.DataFrame(results)


# ══ منتجات مفقودة ════════════════════════════
def find_missing(our_df, comp_dfs):
    our_name_col = best_col(our_df, ["المنتج", "اسم المنتج", "Product", "Name", "name"])
    our_norms    = [normalize(str(r.get(our_name_col, "")))
                    for _, r in our_df.iterrows()
                    if not is_sample(str(r.get(our_name_col, "")))]

    missing, seen = [], set()
    for cname, cdf in comp_dfs.items():
        cn_col = best_col(cdf, ["المنتج", "اسم المنتج", "Product", "Name", "name"])
        ci_col = best_col(cdf, ["ID", "id", "معرف", "SKU", "sku", "الكود", "code"])
        for _, row in cdf.iterrows():
            cp = str(row.get(cn_col, "")).strip()
            if not cp or is_sample(cp): continue
            cn = normalize(cp)
            if not cn or cn in seen: continue
            m = rf_process.extractOne(cn, our_norms, scorer=fuzz.token_sort_ratio, score_cutoff=70)
            if m: continue
            seen.add(cn)
            sz = extract_size(cp)
            missing.append({
                "منتج المنافس": cp,
                "معرف المنافس": get_id(row, ci_col),
                "سعر المنافس":  get_price(row),
                "المنافس":       cname,
                "الماركة":       extract_brand(cp),
                "الحجم":         f"{int(sz)}ml" if sz else "",
                "النوع":         extract_type(cp),
            })
    return pd.DataFrame(missing) if missing else pd.DataFrame()


# ══ تصدير Excel ملوّن ════════════════════════
def export_excel(df, sheet="النتائج"):
    """تصدير مع ألوان حسب القرار"""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    output = io.BytesIO()
    edf = df.copy()
    for c in ["جميع_المرشحين", "جميع المرشحين"]:
        if c in edf.columns: edf.drop(columns=[c], inplace=True)
    sheet_name = str(sheet)[:31]
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        edf.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        hf    = PatternFill("solid", fgColor="1e293b")
        hfont = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[1]:
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
        COLORS = {
            "🔴": "FFE5E5", "🟢": "E5FFE5",
            "✅": "E5FFF5", "⚠️": "FFF8E5", "🔵": "E5F0FF"
        }
        dcol = None
        for i, cell in enumerate(ws[1], 1):
            if "القرار" in str(cell.value or ""): dcol = i; break
        if dcol:
            for ri in range(2, ws.max_row + 1):
                val = str(ws.cell(ri, dcol).value or "")
                for emoji, color in COLORS.items():
                    if emoji in val:
                        for ci in range(1, ws.max_column + 1):
                            ws.cell(ri, ci).fill = PatternFill("solid", fgColor=color)
                        break
        for ci, col in enumerate(ws.columns, 1):
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 55)
    return output.getvalue()
